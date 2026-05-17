"""Read and extract content from existing XLSX files."""

from pathlib import Path
from openpyxl import load_workbook


def read_xlsx(file_path: str) -> str:
    """Extract all text from an XLSX file (all sheets)."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    lines = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"--- Sheet: {sheet_name} ---")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                lines.append("\t".join(cells))
    wb.close()
    return "\n".join(lines)


def get_xlsx_info(file_path: str) -> dict:
    """Get metadata and statistics from an XLSX file."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheets_info = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheets_info.append({
            "name": sheet_name,
            "rows": ws.max_row or 0,
            "columns": ws.max_column or 0,
        })
    wb.close()

    return {
        "file": Path(file_path).name,
        "sheets": len(sheets_info),
        "sheet_details": sheets_info,
    }


def get_xlsx_sheets(file_path: str, sheet_name: str | None = None, max_rows: int = 100) -> list[dict]:
    """Get sheet content as list of rows. Optionally filter by sheet name."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    target_sheets = [sheet_name] if sheet_name else wb.sheetnames
    result = []

    for name in target_sheets:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append([str(c) if c is not None else "" for c in row])
        result.append({
            "sheet": name,
            "rows_returned": len(rows),
            "total_rows": ws.max_row or 0,
            "data": rows,
        })
    wb.close()
    return result
